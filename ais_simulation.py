from turtle import color
from warnings import catch_warnings
import numpy as np
from neuron import h
import pylab as p
from matplotlib import pyplot as plt
import pandas as pd
import openpyxl as op

# Import variables
from ais_variables import *
from ais_plotter import *

# Append to path
import sys
import os
import tables as tbl
import time
import h5py as h5

sys.path.append('..dorsal_horn_network_project/cells')
sys.path.append('..dorsal_horn_network_project/cells/cell_data/numerical_data')

class SimulationDetails(tbl.IsDescription):
    neuron     = tbl.StringCol(10)
    dendL      = tbl.Float64Col()
    spaceL     = tbl.Float64Col()
    Rin        = tbl.Float64Col()
    DensityK   = tbl.Float64Col()
    DensityNa  = tbl.Float64Col()
    globalRa   = tbl.Float64Col()   

class Simulation():
    def __init__(self):
        # Define variables
        self.id = 0
        self.rheobase = None
        self.ap_height = None
        
        # Initialize variables
        self.init_simulation()
    
    def init_simulation(self, vinit=V_INIT, tstop=TSTOP, celsius=CELSIUS):
        """Initialize and run a simulation.
        :param celsius:
        :param v_init:
        :param tstop: Duration of the simulation.
        """
        h.v_init = vinit
        h.tstop = tstop
        h.celsius = celsius
        h.dt = DT
    
    def makeRecorders(self, segment, labels, rec=None):
        if rec is None:
            rec = {'t': h.Vector()}
            rec['t'].record(h._ref_t)
        for k,v in labels.items():
            rec[k] = h.Vector()
            rec[k].record(getattr(segment, v))
        return rec

    def makeIclamp(self, segment, dur, amp, delay=0):
        stim = h.IClamp(segment)
        stim.delay = delay
        stim.dur = dur
        stim.amp = amp
        return stim

    def rheobase_counter(self, cell, min_current=MIN_CURRENT, max_current=MAX_CURRENT, rheobase_step=RHEOBASE_STEP):
        # Remove writing to file and put it seperately
        self.id += 1
        apc = h.APCount(cell.soma(0.5))  
        apc.thresh = THRESHOLD
        ap_badcell = h.APCount(cell.soma(0.5))
        ap_badcell.thresh = -50
        current_lst = np.arange(min_current, max_current, rheobase_step)
        
        # Loop over currents until spike
        for new_current in current_lst:
            cell.stim.amp = new_current
            h.run()
            if apc.n > 0:
                self.ap_height = self.get_ap_amplitude(cell, apc.time)
                self.rheobase = new_current

                # self.ap_amplitude = 
                print("Received {} Spikes with {:2f} pA Rheobase, at time {:2f}".format(apc.n, self.rheobase*1e3, apc.time))
                # print("Received " + str(apc.n) + " Spikes and " + str(self.rheobase*1e3) + " pA Rheobase, at time " + str(apc.time))
                break
            
            if ap_badcell.n > 0:
                self.rheobase = new_current
                break
    
    def get_normalized_data(self, data):
        min_val = min(data)
        max_val = max(data)
        # data = [(x - min_val)/(max_val - min_val) for x in data]
        data  =  [(x - min_val)/(max_val - min_val) for x in data]
        return data
   
    
    def get_voltage_trace(self, cell, plot_flag = True):
        # Initialize APCounter
        apc = h.APCount(cell.soma(RECORDING_LOCATION))
        apc.thresh = THRESHOLD             
        
        # Initialize lists
        time_data = []
        voltage_data = []
        ap_data = []
        for_excel_list = []
        header_names = []
        current_injection_protocol = CUR_INJ_PROTOCOL # Injection current in nA
        
        for new_current in current_injection_protocol:
            cell.stim.amp = new_current 
            h.run()

            # Do it without and with flatten
            ttrace = np.array(cell.t_vec)
            vtrace = np.array(cell.soma_v_vec)
            time_data.append(ttrace)
            voltage_data.append(vtrace)
            
            # Action potential calculation
            ap_n = apc.n / (DUR/1000)
            ap_data.append(ap_n)

            # Save to excel
            header_names.append("Voltage array: {} Hz {} nA".format(ap_n,new_current)) 
            header_names.append("Time array: {} Hz {} nA".format(ap_n,new_current)) 
            for_excel_list.append(ttrace)
            for_excel_list.append(vtrace)
            
            
        # Save into excel
        sheetnamestr = 'Voltage_trace'
        self.save_into_excel(cell, for_excel_list , header_names ,sheetnamestr)

        # Setup Graphs:
        if plot_flag:
            inject_amount = len(current_injection_protocol)
            fig = plt.figure(figsize=(8, 20))
            for i in range(inject_amount):
                ax = fig.add_subplot(inject_amount,1, i+1)
                ax.plot(time_data[i], voltage_data[i], label=str(current_injection_protocol[i]*1e3) + ' pA, Spike Number = {}'.format(ap_data[i]))
                # Format graph
                plt.suptitle('{} Voltage trace'.format(cell.label), fontsize=14, fontweight='bold')
                plt.xlabel('time (ms)')
                # plt.ylabel('mV')
                plt.yticks([])
                ax.axhline(0, color='r', linestyle='dashed')
                ax.legend(prop={'size': 8})
                frame = plt.gca()
                frame.spines['top'].set_visible(False)
                frame.spines['right'].set_visible(False)
                frame.spines['left'].set_visible(False)   
                            
            plt.savefig('cell_data/figures/' + cell.label + '_voltage_trace.svg', format = 'svg', bbox_inches = 'tight', dpi = 300)
            plt.savefig('cell_data/figures/' + cell.label + '_voltage_trace.png', format = 'png', bbox_inches = 'tight', dpi = 300)
            plt.show()  


        ## SMALL Protocol
        # Initialize lists
        time_data = []
        voltage_data = []
        ap_data = []
        for_excel_list = []
        header_names = []
        current_injection_protocol = CUR_INJ_PROTOCOL_SMALL # Injection current in nA

        for new_current in current_injection_protocol:
            cell.stim.amp = new_current 
            h.run()

            # Do it without and with flatten
            ttrace = np.array(cell.t_vec)
            vtrace = np.array(cell.soma_v_vec)
            time_data.append(ttrace)
            voltage_data.append(vtrace)
            
            # Action potential calculation
            ap_n = apc.n / (DUR/1000)
            ap_data.append(ap_n)

            # Save to excel
            header_names.append("Voltage array: {} Hz {} nA".format(ap_n, new_current)) 
            header_names.append("Time array: {} Hz {} nA".format(ap_n, new_current)) 
            for_excel_list.append(ttrace)
            for_excel_list.append(vtrace)
            
        # Save into excel
        sheetnamestr = 'Voltage_trace_small'
        self.save_into_excel(cell, for_excel_list , header_names ,sheetnamestr)

        # Setup Graphs:
        if plot_flag:
            inject_amount = len(current_injection_protocol)
            fig = plt.figure(figsize=(8, 20))
            for i in range(inject_amount):
                ax = fig.add_subplot(inject_amount,1, i+1)
                ax.plot(time_data[i], voltage_data[i], label=str(current_injection_protocol[i]*1e3) + ' pA, Spike Number = {}'.format(ap_data[i]))
                # Format graph
                plt.suptitle('{} Voltage trace'.format(cell.label), fontsize=14, fontweight='bold')
                plt.xlabel('time (ms)')
                # plt.ylabel('mV')
                plt.yticks([])
                ax.axhline(0, color='r', linestyle='dashed')
                ax.legend(prop={'size': 8})
                frame = plt.gca()
                frame.spines['top'].set_visible(False)
                frame.spines['right'].set_visible(False)
                frame.spines['left'].set_visible(False)   
                            
            plt.savefig('cell_data/figures/' + cell.label + '_voltage_trace_small.svg', format = 'svg', bbox_inches = 'tight', dpi = 300)
            plt.savefig('cell_data/figures/' + cell.label + '_voltage_trace_small.png', format = 'png', bbox_inches = 'tight', dpi = 300)
            plt.show()    


    def get_if(self, cell, delay = DELAY, dur = DUR, plot_f = True):
        """
        Extract the frequency output curve of the cell
        """
        segment = cell.soma(0.5)
        aps = []
        
        # Instill action potential counter
        ap = h.APCount(segment)
        
        # IF Injection currents
        cur_inj_range = np.linspace(0, 0.13, 13)

        # Loop over injection
        for inj in cur_inj_range:
            cell.stim.amp = inj
            h.run()

            #Number of action potentials - Divided by sampling time
            ap_n = ap.n / (DUR/1000)
            aps.append(ap_n)
        
        # Save to excel:
        list_of_elem = [cur_inj_range, aps]
        headers = ['Current (nA)', 'Frequency (Hz)']
        sheetnamestr = 'FI_Curve'
        self.save_into_excel(cell, list_of_elem , headers ,sheetnamestr)
 
        if plot_f:
            self.plot_if(cur_inj_range, aps, cell)        

    def plot_if(self, current_vector, freq_vector, cell):
        """
        Description: Plots the IF curves of 
        """        
        plt.plot(current_vector, freq_vector ,marker='o',color='black', linewidth=2)
        # ax1.plot(current_vector, freq_onset_vector, '--', color=color_vec[1][0], label = currlabel + " onset rate")
        plt.xlabel("Current [nA]")
        plt.ylabel("Frequency [Hz]")
        plt.title("I/F")
        lg = plt.legend()
        lg.get_frame().set_linewidth(0.5)
        plt.savefig('cell_data/figures/' + cell.label + 'IF_Curve.svg', format = 'svg', bbox_inches = 'tight', dpi = 300)
        plt.savefig('cell_data/figures/' + cell.label + 'IF_Curve.png', format = 'png', bbox_inches = 'tight', dpi = 300)
        plt.show()
        

    def get_channel_conductances(self, cell):
        segment = cell.AIS(0.5)
        stim = self.makeIclamp(segment, DUR, 0, DELAY)
        rec_na = self.makeRecorders(segment, {'v': '_ref_ina'})
        rec_k = self.makeRecorders(segment, {'v': '_ref_ik'})
        stim.amp = 0.06
        h.run()
        t = np.array(rec_na['t'])
        # Recording of voltage in mV
        v_na = np.array(rec_na['v'])
        v_k = np.array(rec_k['v'])
        p.plot(1e-3*t,v_na)
        p.plot(1e-3*t,v_k)
        p.xlabel('t (s)')
        p.ylabel('V (mV)')
        p.show()
    
    def get_ap_amplitude(self, cell, ap_time):
        """
        Get amplitude of action potential
        Achieves so getting action potential 100 ms within first threshold cross
        Inputs:
         - cell - Neuron model with all characteristics
         - ap_time - The time of the action potential
        """
        v_np_array = np.array(cell.soma_v_vec)
        t_np_array = np.array(cell.t_vec)
        ap_index_start = int(ap_time / DT)      #
        ap_index_end = ap_index_start + int(50/DT) # + 50 ms range
        ap_height = max(v_np_array[ap_index_start:ap_index_end])
        # print('ap_time: ',ap_time)
        # print('ap_time: ',t_np_array[ap_index_start])
        print('The AP height is {} mV'.format(ap_height))
        return ap_height

    def get_input_resistance(self, cell, delay=DELAY, dur=DUR, plot_flag=True, save_to_excel = False):
        """
        Description: Extracts the input resistance from the neuron 

        The protocol from the experiment was:
        (RIN) was measured in voltage-clamp mode using negative 10- to 40-mV pulses from a holding level of –80 mV. 
        Only cells with a resting potential (VR) negative to –60 mV were included into this study RIN of 1.7 GΩ
        
        For the original mode the Rin was assumed to be around 1.7G, According to recorded cell which were 1.7 ± 0.3 GΩ
        """
        
        # Define variables
        segment = cell.soma(0.5)
        current_arr = [-0.3,-0.12,0.04]
        rec = self.makeRecorders(segment, {'v': '_ref_v'})
        ap = h.APCount(segment)
        ap.thresh = -20
        spks = h.Vector()
        ap.record(spks)
        I = []
        V = []
        
        # Variables to save for excel
        headers = []
        list_of_elem = []
        list_of_elem_iv = []
        
        if plot_flag:
            p.figure()
            p.subplot(1,2,1)
        for k,i in enumerate(np.arange(current_arr[0],current_arr[1],current_arr[2])):     
            spks.clear()
            ap.n = 0
            cell.stim.amp = i
            h.run()
            spike_times = np.array(spks)
            if len(np.intersect1d(np.nonzero(spike_times>delay)[0], np.nonzero(spike_times<delay+dur)[0])) == 0:
                # Recording of time calculated in ms   
                t = np.array(rec['t'])
                # Recording of voltage in mV
                v = np.array(rec['v'])
                # Extract steady state of the voltage
                idx = np.intersect1d(np.nonzero(t > delay+0.75*dur)[0], np.nonzero(t < delay+dur)[0])
                # Insert current
                I.append(i)
                # Calculate the mean of the steady state state, and substract the resting voltage pot
                V.append(np.mean(v[idx]) + 70) 
            else:
                print('The neuron emitted spikes at I = %g pA' % (cell.stim.amp*1e3))
            
            # Convert to different units:
            t_new_units = 1e-3*t  # What units is this?

            # Save data to excel
            if save_to_excel:
                list_of_elem.append(t_new_units)
                list_of_elem.append(v)
                headers.append('Time Trace (s): {} pA inj'.format(i))
                headers.append('Voltage Trace(mV): {} pA inj'.format(i))
            
            if plot_flag:
                p.plot(1e-3*t,v)
            
        #? Covert to microvolt, why ar we doing this? to help with the fit?        
        V = np.array(V)*1e-3
        # Convert current to pA units
        I = np.array(I)*1e-9

        #? Verify the polyfit function
        poly = np.polyfit(I,V,1)
        if plot_flag:
            # Format the plot and plot the results
            ymin,ymax = p.ylim()
            p.plot([1e-3*(delay+0.75*dur),1e-3*(delay+0.75*dur)],[ymin,ymax],'r--')
            p.plot([1e-3*(delay+dur),1e-3*(delay+dur)],[ymin,ymax],'r--')
            p.xlabel('t (s)')
            p.ylabel('V (mV)')
            p.box(True)
            p.grid(False)
            p.subplot(1,2,2)
            
            # Plots the current injected
            x = np.linspace(I[0],I[-1],100)
            y = np.polyval(poly,x)
            p.plot(1e12*x,1e3*y,'k--')
            p.plot(1e12*I,1e3*V,'bo')
            p.xlabel('I (pA)')
            p.ylabel('V (mV)')
            # Save the figures
            plt.savefig('cell_data/figures/' + cell.label + '_IV_Rin_Protocol.svg', format = 'svg', bbox_inches = 'tight', dpi = 1200)
            plt.savefig('cell_data/figures/' + cell.label + '_IV_Rin_Protocol.png', format = 'png', bbox_inches = 'tight', dpi = 1200)
            p.show()       
        
        #Convert to MegaOhm
        Rin = poly[0]*1e-6
        
        # Save the data
        # np.save('cell_data/figures/IV_data', [V, I])
        print('The cell input resistance is ' + str(Rin))

        # Save the data to excel
        # Save IV curve
        if save_to_excel:
            sheetnamestr_iv = 'IV Curve'
            headers_iv = ['I (pA)', 'V (mV)']
            list_of_elem_iv.append(1e12*I)
            list_of_elem_iv.append(1e3*V)
            self.save_into_excel(cell, list_of_elem_iv , headers_iv ,sheetnamestr_iv)
        
        # Save voltage traces
        if save_to_excel:
            sheetnamestr = 'Input Resistance'
            self.save_into_excel(cell, list_of_elem , headers ,sheetnamestr)

        # Return the data
        return Rin
    
    def get_phase_plane_trace(self, cell):
        ais_list = [5, 60]
        # Make stim object       
        # Plot the phase plane of the graphs
        fig1 = p.figure(1)      
        fig1_1 = fig1.add_subplot(221)
        fig1_2 = fig1.add_subplot(223)
        colors=["orangered","darkred","gold"]
        for index, item in enumerate(ais_list):
            cell.spacer.L = item
            self.rheobase_protocol(cell)
            cell.stim.amp = self.rheobase
            print("Iinj =", self.rheobase, "nA")
            h.run()    
            time = np.array(cell.t_vec)
            vtrace=np.array(cell.soma_v_vec).flatten()
            # sliced_vtrace = vtrace[round(740/DT):round(760/DT)]
            # sliced_time = time[round(740/DT):round(760/DT)]
            
            # vtraceAIS=np.array(cell.AIS_v_vec).flatten()
            dv= self.extract_dv_dt(vtrace)
            # dvAIS= self.extract_dv_dt(vtraceAIS) 
            # Plot the data 
            fig1_1.plot(time, vtrace, color=colors[index])
            # fig1_1.plot(time, vtraceAIS, color =colors[index], linestyle="dashed")     
            fig1_2.plot(vtrace[: len(dv)], dv, color=colors[index])
            # fig1_2.plot(vtraceAIS[: len(dvAIS)], dvAIS,linestyle="dashed", color=colors[index])   
    
        # Edit the files
        fsize = 10
        fig1_1.set_xlabel("Time [ms]", fontsize=fsize)
        fig1_1.set_ylabel("Voltage [mV]", fontsize=fsize)
        # fig1_1.set_xlim([700,800])
        fig1_2.set_xlabel("Voltage [mV]", fontsize=fsize)    
        fig1_2.set_ylabel("dV/dt [V/s]", fontsize=fsize)
        # Format the graph
        #format the plot    
        p.figure(1)
        # fig1_1.set_title("Iinj = 0.4 nA (red), 0.8 nA (red), 1.3 nA (orange) \n Soma (solid) and AIS (dashed)")
        fig1_1.set_title("Bright red {} AIS dark red {} AIS".format(5,60))
        plt.savefig('cell_data/figures/' + cell.label + 'Phase_Plane_Dynamics.svg', format = 'svg', bbox_inches = 'tight', dpi = 300)
        plt.savefig('cell_data/figures/' + cell.label + 'Phase_Plane_Dynamics.png', format = 'png', bbox_inches = 'tight', dpi = 300)
        p.show()
    
    def rheobase_protocol(self, cell, plot_flag = False):
        apc = h.APCount(cell.soma(0.5))
        apc.thresh = 0
        current_lst = np.arange(MIN_CURRENT, MAX_CURRENT, RHEOBASE_STEP)
        for new_current in current_lst:
            cell.stim.amp = new_current
            h.run()
            if apc.n > 0:
                self.ap_height = self.get_ap_amplitude(cell, apc.time)
                self.rheobase = new_current
                if plot_flag:
                    # Plot the results
                    plt.plot(cell.t_vec, cell.soma_v_vec, label=str(new_current) + ', Rheobase = {}'.format(new_current))
                    plt.suptitle('Spike Graph', fontsize=14, fontweight='bold')
                    plt.axvline(x = apc.time, color = 'r')
                    plt.axhline(y = self.ap_height, color = 'b', linestyle = '-')
                    # plt.text(0.1, 2.8, "The number of action potentials is {}".format(apc.n))
                    plt.xlabel('time (ms)')
                    plt.ylabel('mV')
                    plt.xlim(600,700)
                    plt.legend()
                    plt.show()
                    # Get ap height:
      
                break
    
    def get_ais_distalization_effect(self, cell, plot_flag = True):
        # Initialize arrays
        time_arr = []
        soma_v_arr = []
        rheobase_arr = []
        Rin_arr = []
        Vth_arr = []
        ap_arr = []     
        
        for index, item in enumerate(SPACER_ARR):
            # Set up simulation
            cell.spacer.L = item
            
            # Get rheobase for current ais distance
            Rin = self.get_input_resistance(cell, plot_flag=False)
            self.rheobase_counter(cell)
            
            # Append values
            rheobase_arr.append(self.rheobase)
            time_arr.append(np.array(cell.t_vec))
            soma_v_arr.append(np.array(cell.soma_v_vec))
            ap_arr.append(self.ap_height)
            Rin_arr.append(Rin)         
        
        # Save the data into an h5 file
        output_file_name = self.create_output_filename('cell_data/numerical_data/'+ cell.__repr__()+ 'ais_plasticity_', '.h5')
        
        # Save into excel
        header_names = ['AIS Distance (um)','Rheobase (nA)', 'Input Resistance (MegaOhm)', 'Voltage threshold (mV)', 'AP Height (mV)']
        for_excel_list = [SPACER_ARR,rheobase_arr, Rin_arr, Vth_arr, ap_arr]
        sheetnamestr = 'Excitability Correlation'
        self.save_into_excel(cell, for_excel_list , header_names ,sheetnamestr)
        
        # Plot and save the data:
        if plot_flag == True:
            my_plotter(cell, time_arr, soma_v_arr, SPACER_ARR, rheobase_arr)
            correlation_plotter(cell, Rin_arr, SPACER_ARR, header_names[1])
            correlation_plotter(cell, ap_arr, SPACER_ARR, header_names[3])

        return rheobase_arr

    def get_plasticity_change_multi(self, cell, param):
        header_names = []
        for_excel_list_rheobase_mat = []
        value = PARAM_SPACE[param]
        parameter_range = np.linspace(value[0], value[1], value[2]) 

        # Consider only the last two values
        parameter_range = [float('%.2f' % (f)) for f in parameter_range] 
        length = value[2]
        for index, param_val in enumerate(parameter_range):
            # Example: cell.dend.L is the param 
            exec(param + '=' + str(param_val))
            rheobase_arr = self.get_ais_distalization_effect(cell)
            print("Rheobase results ready for " + param + " = " + str(param_val))
            print(str((index/length)*100)+ "% Simulation Complete")
            
            # result = np.polyfit(list_of_index, list(array_of_data), order)
            # Save to excel
            header_names.append("{} parameter: {} length".format(param,param_val)) 
            for_excel_list_rheobase_mat.append(rheobase_arr)
            
        # Save into excel:
        sheetnamestr = 'Multi Excitability Correlation'
        self.save_into_excel(cell, for_excel_list_rheobase_mat , header_names ,sheetnamestr)

        # output_file_name = param + 'simulation'
        # np.savez(output_file_name, space_arr = SPACER_ARR, parameter_range = parameter_range, parameter = value, rheobase_mat = rheobase_mat)
        # print('Succesfully saved simulation matrix for' + param)
        # data = np.load(output_file_name)
        # return rheobase_mat
        # plot_ais_plasticity_change(self, cell)

    def extract_dv_dt(self, vtrace):
        '''2-point first order finite difference to estimate dV/dt '''
        dt = DT
        dv = []
        for i in range(1, len(vtrace)-2): 
            dv.append((vtrace[i+1]-vtrace[i-1])/(2*dt))
            # dv.append((vtrace[i+1]-vtrace[i-1])/(dt))
        return dv

    def create_output_filename(self, prefix='', extension='.h5'):
        """
        Create output filename for saving dat
        """
        filename = prefix
        if prefix != '' and prefix[-1] != '_':
            filename = filename + '_'
        now = time.localtime(time.time())
        filename = filename + '%d%02d%02d-%02d%02d%02d' % \
            (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
        if extension[0] != '.':
            extension = '.' + extension
        suffix = ''
        k = 0
        while os.path.exists(filename + suffix + extension):
            k = k+1
            suffix = '_%d' % k
        return filename + suffix + extension

    def save_simulation(self, filename, neuron_type, dendL, diam, Rin, model_gk, model_Na, Ra, arr_rheobase, arr_spacer):
        ais_file = tbl.open_file(filename, mode='w', title='AIS Plasticity Simulations')
        table = ais_file.create_table(ais_file.root, 'Details', SimulationDetails, 'Simulation info')
        details = table.row
        details['neuron'] = neuron_type
        details['dendL'] = dendL
        details['spaceL'] = diam
        details['Rin'] = Rin
        details['DensityK'] = model_gk
        details['DensityNa'] = model_Na
        details['globalRa'] = Ra
        details.append()
        group = ais_file.create_group(ais_file.root, 'Data', 'Rheobase and AIS to Soma Distances')
        ais_file.create_array(group, 'Rheobase', arr_rheobase, 'Rheobase measurements')
        ais_file.create_array(group, 'Spacers', arr_spacer, 'Spacer lengths')
        ais_file.close()
        

    def load_simulation(self, filename): 
        with h5.File(filename, "r") as f:  
            # List all groups
            print("Keys: %s" % f.keys())
            simulation_labels = list(f.keys())[1]
            simulation_data = list(f[simulation_labels])
            print(simulation_data)
            a_group_key = list(f.keys())[0]
            # Get the data
            rheobase_data = list(f[a_group_key]['Rheobase'])
            spacers_data = list(f[a_group_key]['Spacers'])
        return rheobase_data

    def save_into_excel(self, laminaNeuron, list_of_elem , header_names ,sheetnamestr):
        neuron_characteristics = [self.id, 1 / laminaNeuron.dend.g_pas,
                          laminaNeuron.dend.Ra, laminaNeuron.dend.cm, laminaNeuron.dend.L,
                          laminaNeuron.spacer.L, 1 / laminaNeuron.spacer.g_pas, laminaNeuron.spacer.Ra,
                          laminaNeuron.spacer.cm, 1 / laminaNeuron.AIS.g_pas, self.rheobase]
        
        # Define the writer and book
        try:
            book = op.load_workbook('cell_data/numerical_data/{}.xlsx'.format(laminaNeuron.label))
        except:
            book = op.Workbook()
            # ws = book.active
            # ws.title = sheetnamestr
            book.save(filename='cell_data/numerical_data/{}.xlsx'.format(laminaNeuron.label))
        
        writer = pd.ExcelWriter('cell_data/numerical_data/{}.xlsx'.format(laminaNeuron.label), engine='openpyxl')
        writer.book = book
        
        # Define the dataframe
        df = pd.DataFrame(list_of_elem).T
        df.columns=header_names

        # Convert to excel
        df.to_excel(writer, sheet_name=sheetnamestr, index=False)
        # worksheet2 = workbook.add_worksheet('Data') 
        # worksheet = writer.sheets[sheetnamestr]
        # worksheet.set_column(1, 3, len(header_names[0]))  # Width of columns B:D set to 30.
        writer.save()
        writer.close()
        
        